"""
Service d'export - Excel et PDF
Export des opportunités filtrées avec mise en forme professionnelle
"""
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
from io import BytesIO
import base64

from sqlalchemy.orm import Session

from app.db.models.opportunity import Opportunity, OpportunityStatus, OpportunityCategory

logger = logging.getLogger(__name__)


class ExportService:
    """Service d'export des données"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def export_opportunities_excel(
        self,
        opportunities: List[Opportunity],
        include_fields: Optional[List[str]] = None
    ) -> bytes:
        """
        Exporte les opportunités au format Excel
        
        Args:
            opportunities: Liste des opportunités à exporter
            include_fields: Champs à inclure (tous si None)
        
        Returns:
            Contenu du fichier Excel en bytes
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise ImportError("openpyxl required for Excel export. Install with: pip install openpyxl")
        
        # Créer le workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Opportunités"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Colonnes par défaut
        default_columns = [
            ("ID", 10),
            ("Titre", 50),
            ("Organisation", 30),
            ("Catégorie", 20),
            ("Status", 15),
            ("Score", 10),
            ("Budget (€)", 15),
            ("Deadline", 15),
            ("Région", 20),
            ("Source", 20),
            ("URL", 40),
            ("Créé le", 15),
        ]
        
        # En-têtes
        for col, (header, width) in enumerate(default_columns, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Données
        for row, opp in enumerate(opportunities, start=2):
            data = [
                opp.id,
                opp.title[:100] if opp.title else "",
                opp.organization or "",
                opp.category.value if opp.category else "",
                opp.status.value if opp.status else "",
                opp.score or 0,
                float(opp.budget_amount) if opp.budget_amount else "",
                opp.deadline_at.strftime("%d/%m/%Y") if opp.deadline_at else "",
                opp.location_region or "",
                opp.source_name or "",
                opp.url_primary or "",
                opp.created_at.strftime("%d/%m/%Y") if opp.created_at else "",
            ]
            
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                
                # Couleur conditionnelle pour le score
                if col == 6 and isinstance(value, (int, float)):  # Score
                    if value >= 15:
                        cell.fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
                    elif value >= 10:
                        cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
                    elif value < 5:
                        cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
                
                # Couleur conditionnelle pour le status
                if col == 5:  # Status
                    status_colors = {
                        "WON": "22C55E",
                        "LOST": "EF4444",
                        "NEW": "3B82F6",
                        "IN_PROGRESS": "F59E0B",
                    }
                    if value in status_colors:
                        cell.fill = PatternFill(
                            start_color=status_colors[value],
                            end_color=status_colors[value],
                            fill_type="solid"
                        )
                        cell.font = Font(color="FFFFFF")
        
        # Filtres auto
        ws.auto_filter.ref = f"A1:L{len(opportunities) + 1}"
        
        # Figer la première ligne
        ws.freeze_panes = "A2"
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def export_opportunities_csv(
        self,
        opportunities: List[Opportunity],
        delimiter: str = ";"
    ) -> str:
        """
        Exporte les opportunités au format CSV
        
        Args:
            opportunities: Liste des opportunités
            delimiter: Séparateur (défaut: ; pour Excel FR)
        
        Returns:
            Contenu CSV en string
        """
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
        
        # En-têtes
        headers = [
            "ID", "Titre", "Organisation", "Catégorie", "Status", "Score",
            "Budget", "Deadline", "Région", "Ville", "Source", "URL", "Créé le"
        ]
        writer.writerow(headers)
        
        # Données
        for opp in opportunities:
            row = [
                opp.id,
                opp.title or "",
                opp.organization or "",
                opp.category.value if opp.category else "",
                opp.status.value if opp.status else "",
                opp.score or 0,
                float(opp.budget_amount) if opp.budget_amount else "",
                opp.deadline_at.strftime("%d/%m/%Y") if opp.deadline_at else "",
                opp.location_region or "",
                opp.location_city or "",
                opp.source_name or "",
                opp.url_primary or "",
                opp.created_at.strftime("%d/%m/%Y %H:%M") if opp.created_at else "",
            ]
            writer.writerow(row)
        
        return output.getvalue()
    
    def export_opportunities_pdf(
        self,
        opportunities: List[Opportunity],
        title: str = "Rapport des Opportunités",
        include_stats: bool = True
    ) -> bytes:
        """
        Exporte les opportunités au format PDF
        
        Args:
            opportunities: Liste des opportunités
            title: Titre du rapport
            include_stats: Inclure les statistiques en haut
        
        Returns:
            Contenu PDF en bytes
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            raise ImportError("reportlab required for PDF export. Install with: pip install reportlab")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=15*mm,
            leftMargin=15*mm,
            topMargin=15*mm,
            bottomMargin=15*mm
        )
        
        styles = getSampleStyleSheet()
        elements = []
        
        # Titre
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=12,
            textColor=colors.HexColor('#6366F1'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(
            f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
            ParagraphStyle('Subtitle', parent=styles['Normal'], alignment=TA_CENTER, textColor=colors.gray)
        ))
        elements.append(Spacer(1, 10*mm))
        
        # Statistiques
        if include_stats and opportunities:
            stats_data = self._calculate_stats(opportunities)
            stats_table_data = [
                ["Total", "Nouvelles", "Qualifiées", "Gagnées", "Budget Total", "Score Moyen"],
                [
                    str(stats_data['total']),
                    str(stats_data['new']),
                    str(stats_data['qualified']),
                    str(stats_data['won']),
                    f"{stats_data['total_budget']:,.0f} €",
                    f"{stats_data['avg_score']:.1f}/20"
                ]
            ]
            
            stats_table = Table(stats_table_data, colWidths=[40*mm]*6)
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#F3F4F6')),
                ('FONTSIZE', (0, 1), (-1, 1), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ]))
            elements.append(stats_table)
            elements.append(Spacer(1, 10*mm))
        
        # Table des opportunités
        table_data = [["ID", "Titre", "Organisation", "Status", "Score", "Budget", "Deadline"]]
        
        for opp in opportunities[:100]:  # Limiter à 100 pour éviter les PDF trop longs
            table_data.append([
                str(opp.id),
                (opp.title[:40] + "...") if len(opp.title or "") > 40 else (opp.title or ""),
                (opp.organization[:25] + "...") if len(opp.organization or "") > 25 else (opp.organization or "N/A"),
                opp.status.value if opp.status else "",
                f"{opp.score}/20" if opp.score else "N/A",
                f"{float(opp.budget_amount):,.0f} €" if opp.budget_amount else "N/A",
                opp.deadline_at.strftime("%d/%m/%Y") if opp.deadline_at else "N/A"
            ])
        
        col_widths = [15*mm, 70*mm, 50*mm, 30*mm, 20*mm, 30*mm, 25*mm]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Style de la table
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # ID centré
            ('ALIGN', (3, 1), (-1, -1), 'CENTER'),  # Status, Score, Budget, Deadline centrés
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ]
        
        # Couleurs conditionnelles pour les status
        for i, opp in enumerate(opportunities[:100], start=1):
            if opp.status == OpportunityStatus.WON:
                style_commands.append(('BACKGROUND', (3, i), (3, i), colors.HexColor('#22C55E')))
                style_commands.append(('TEXTCOLOR', (3, i), (3, i), colors.white))
            elif opp.status == OpportunityStatus.LOST:
                style_commands.append(('BACKGROUND', (3, i), (3, i), colors.HexColor('#EF4444')))
                style_commands.append(('TEXTCOLOR', (3, i), (3, i), colors.white))
        
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        
        # Footer
        if len(opportunities) > 100:
            elements.append(Spacer(1, 5*mm))
            elements.append(Paragraph(
                f"... et {len(opportunities) - 100} autres opportunités non affichées",
                ParagraphStyle('Footer', parent=styles['Normal'], alignment=TA_CENTER, textColor=colors.gray)
            ))
        
        doc.build(elements)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def _calculate_stats(self, opportunities: List[Opportunity]) -> Dict[str, Any]:
        """Calcule les statistiques pour un ensemble d'opportunités"""
        total = len(opportunities)
        new = sum(1 for o in opportunities if o.status == OpportunityStatus.NEW)
        qualified = sum(1 for o in opportunities if o.status == OpportunityStatus.QUALIFIED)
        won = sum(1 for o in opportunities if o.status == OpportunityStatus.WON)
        
        budgets = [float(o.budget_amount) for o in opportunities if o.budget_amount]
        total_budget = sum(budgets)
        
        scores = [o.score for o in opportunities if o.score]
        avg_score = sum(scores) / len(scores) if scores else 0
        
        return {
            "total": total,
            "new": new,
            "qualified": qualified,
            "won": won,
            "total_budget": total_budget,
            "avg_score": avg_score
        }
    
    def export_to_base64(self, content: bytes, format: str) -> Dict[str, str]:
        """Encode le contenu en base64 pour téléchargement via API"""
        encoded = base64.b64encode(content).decode('utf-8')
        
        mime_types = {
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "csv": "text/csv",
            "pdf": "application/pdf"
        }
        
        return {
            "content": encoded,
            "mime_type": mime_types.get(format, "application/octet-stream"),
            "format": format
        }
